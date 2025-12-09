"""
Testovací modul pro parsování SPE/CNF spektra a zápis do Excel formátu
"""

import pandas as pd
from openpyxl import Workbook
from scripts.utils import parse_spe_file

def read_cnf_file(filepath):
    """Read CNF file (same format as SPE)"""
    # Try different encodings
    for encoding in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
        try:
            with open(filepath, 'r', encoding=encoding) as f:
                content = f.read()
            print(f"✓ Úspěšně načteno s enkódováním: {encoding}")
            return content
        except UnicodeDecodeError:
            continue
    
    # Fallback - read as binary and decode with errors='ignore'
    with open(filepath, 'r', encoding='latin1', errors='ignore') as f:
        content = f.read()
    print(f"✓ Načteno s latin1 (ignorování chyb)")
    return content

def create_excel_from_multiple_spe(spe_data_list, output_file):
    """Create Excel file with multiple samples in SPE-style format"""
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # === VZORKY SHEET (SPE-style format with multiple samples) ===
    ws_vzorky = wb.create_sheet('Vzorky')
    
    # Determine max channels
    max_channels = max(len(spe['channels']) for spe in spe_data_list)
    
    # Build rows
    # Row 0: Header with sample names
    header_row = [''] + [spe['sample_name'] for spe in spe_data_list]
    ws_vzorky.append(header_row)
    
    # Row 1: $SPEC_ID:
    spec_id_row = ['$SPEC_ID:'] + [spe['sample_name'] for spe in spe_data_list]
    ws_vzorky.append(spec_id_row)
    
    # Row 2: $MEAS_TIM:
    meas_tim_row = ['$MEAS_TIM:'] + [spe.get('ELIVE', 0) for spe in spe_data_list]
    ws_vzorky.append(meas_tim_row)
    
    # Row 3-6: $MCA_CAL:
    ws_vzorky.append(['$MCA_CAL:'] + [''] * len(spe_data_list))
    ws_vzorky.append(['  a0'] + [spe.get('ECOFFSET', 0) for spe in spe_data_list])
    ws_vzorky.append(['  a1'] + [spe.get('ECSLOPE', 1) for spe in spe_data_list])
    ws_vzorky.append(['  a2'] + [spe.get('ECQUAD', 0) for spe in spe_data_list])
    
    # Row 7: $DATA:
    ws_vzorky.append(['$DATA:'] + [''] * len(spe_data_list))
    
    # Channel data rows
    for ch_num in range(max_channels):
        row = [ch_num]
        for spe in spe_data_list:
            counts = spe['channels'][ch_num] if ch_num < len(spe['channels']) else 0
            row.append(counts)
        ws_vzorky.append(row)
    
    # === KALIBRACE SHEET (prázdná pro tento test) ===
    ws_calib = wb.create_sheet('Kalibrace')
    ws_calib.append(['CHNL', 'Ra', 'K', 'Th'])
    
    # === POZADÍ SHEET (prázdná) ===
    ws_bg = wb.create_sheet('Pozadí')
    ws_bg.append(['CHNL', 'BG_test'])
    
    # === PARAMETRY SHEET ===
    ws_params = wb.create_sheet('Parametry')
    # Use first sample's calibration as reference
    ws_params.append(['ref_a0', spe_data_list[0].get('ECOFFSET', 0)])
    ws_params.append(['ref_a1', spe_data_list[0].get('ECSLOPE', 1)])
    ws_params.append(['ref_a2', spe_data_list[0].get('ECQUAD', 0)])
    
    # Save
    wb.save(output_file)
    print(f"✅ Excel vytvořen: {output_file}")
    print(f"   Počet vzorků: {len(spe_data_list)}")
    for spe in spe_data_list:
        print(f"   - {spe['sample_name']}: {len(spe['channels'])} kanálů, ELIVE={spe.get('ELIVE', 0)}")
    print(f"\n📊 Formát: SPE-style s {len(spe_data_list)} vzorky ve sloupcích")

if __name__ == '__main__':
    # Load multiple SPE files from naitl directory
    import glob
    
    spe_files = glob.glob(r'Spektra\naitl\*.SPE')[:5]  # First 5 files for testing
    output_file = 'Spektra - NaITl - test.xlsx'
    
    print(f"Načítám {len(spe_files)} SPE souborů...")
    
    spe_data_list = []
    
    for input_file in spe_files:
        try:
            print(f"  - {input_file}")
            
            # Read and parse
            spe_content = read_cnf_file(input_file)
            spe_data = parse_spe_file(spe_content)
            
            # Extract sample name from filename
            import os
            sample_name = os.path.splitext(os.path.basename(input_file))[0]
            spe_data['sample_name'] = sample_name
            
            spe_data_list.append(spe_data)
            
        except Exception as e:
            print(f"    ❌ Chyba: {e}")
    
    if spe_data_list:
        # Create Excel
        create_excel_from_multiple_spe(spe_data_list, output_file)
        print("\nHotovo! Zkontrolujte soubor:", output_file)
    else:
        print("Žádné SPE soubory se nepodařilo načíst")
